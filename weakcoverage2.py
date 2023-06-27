#!/usr/bin/env python
# -*- coding:utf-8 -*-

from xlrd import *
from openpyxl import Workbook


def demoTest(provience, num):
    # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
    wb = Workbook()
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    ws1 = wb.active
    ws1.title = "宏站弱覆盖"  # 修改sheet1的名字
    ws2 = wb.create_sheet(title="室分弱覆盖")  # 新建sheet2

    f = open("E:\\weakcoverage\October.txt", "r", encoding='utf-8')  # 在路径以utf-8的格式打开txt文件

    lines = f.readlines()  # read all lines
    sam_point = 0  # 初始化 采样点
    cov_rat = 0  # 覆盖率
    inout_door = 0  # 室内外
    core_city = 0  # 核心城区
    line_1 = lines[0].strip().split('|')  # deal the first line.
    for k in range(len(line_1)):  # get the position of each data
        if line_1[k] == "采样点数量":
            sam_point = k
        if line_1[k] == "覆盖率":
            cov_rat = k
        if line_1[k] == "室内外":
            inout_door = k
        if line_1[k] == "是否核心城区":
            core_city = k
    ws1.append(line_1)
    ws2.append(line_1)

    # file = Workbook()
    # table = file.add_sheet("d:\\dataof8\A123.xlsx")
    j = 1
    z = 1  # start from line 2
    for line in lines[1:len(lines) - 1]:
        p = line.strip().split('|')  # split every line by symbol '|'
        if (p[0] == provience):
            try:
                p[sam_point] = int(p[sam_point])  # transform the type
                p[cov_rat] = float(p[cov_rat])
            except BaseException:  # Maybe show mistake:"list index out of range"
                print(p)
                continue
            # ['FJ', '南平', '邵武余山ZLH1-2', '117.10498', '27.15513', '25196-131', '166266', '0.2573', 'outdoor', '0']
            try:
                if (p[sam_point] >= 1000 and p[cov_rat] <= 0.8) and p[inout_door] == 'outdoor' and p[core_city] == '1':
                    # print(p)
                    for i in range(len(p)):
                        ws1.cell(row=j + 1, column=i + 1).value = p[i]  # 把数据挨个读入sheet1中
                    j = j + 1
            except BaseException:  # Maybe show mistake:"list index out of range"
                print(p)
                continue
                # print(j)
            if (p[sam_point] >= 1000 and p[cov_rat] <= 0.9) and p[inout_door] == 'indoor':
                for i in range(len(p)):
                    ws2.cell(row=z + 1, column=i + 1).value = p[i]
                z = z + 1
        else:
            continue

    wb.save('E:\\weakcoverage\dataof8\%s.xlsx' % provience)

    print('%s is ok' % provience)
    f.close()


def demoTest1(provience, num):
    # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
    wb = Workbook()
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    ws1 = wb.active
    ws1.title = "宏站弱覆盖"
    # from openpyxl import Workbook

    ws2 = wb.create_sheet(title="室分弱覆盖")

    f = open("E:\\weakcoverage\October.txt", "r", encoding='utf-8')

    lines = f.readlines()  # read all lines
    sam_point = 0
    cov_rat = 0
    inout_door = 0
    core_city = 0
    line_1 = lines[0].strip().split('|')  # deal the first line
    for k in range(len(line_1)):  # get the position of each data
        if line_1[k] == "采样点数量":
            sam_point = k
        if line_1[k] == "覆盖率":
            cov_rat = k
        if line_1[k] == "室内外":
            inout_door = k
        if line_1[k] == "是否核心城区":
            core_city = k
    ws1.append(line_1)
    ws2.append(line_1)

    # file = Workbook()
    # table = file.add_sheet("d:\\dataof8\A123.xlsx")
    j = 1
    z = 1  # start from line 2
    for line in lines[1:len(lines) - 1]:
        p = line.strip().split('|')  # split every line by symbol '|'
        if (p[0] != provience):
            try:
                p[sam_point] = int(p[sam_point])  # transform the type
                p[cov_rat] = float(p[cov_rat])
            except BaseException:  # Maybe show mistake:"list index out of range"
                print(p)
                continue
            # ['FJ', '南平', '邵武余山ZLH1-2', '117.10498', '27.15513', '25196-131', '166266', '0.2573', 'outdoor', '0']
            try:
                if (p[sam_point] >= 1000 and p[cov_rat] <= 0.8) and p[inout_door] == 'outdoor' and p[core_city] == '1':
                    # print(p)
                    for i in range(len(p)):
                        ws1.cell(row=j + 1, column=i + 1).value = p[i]
                    j = j + 1
            except BaseException:  # Maybe show mistake:"list index out of range"
                print(p)
                continue
                # print(j)
            if (p[sam_point] >= 1000 and p[cov_rat] <= 0.9) and p[inout_door] == 'indoor':
                for i in range(len(p)):
                    ws2.cell(row=z + 1, column=i + 1).value = p[i]
                z = z + 1
        else:
            continue

    wb.save('E:\\weakcoverage\dataof8\%s.xlsx' % provience)
    print('%s is ok' % provience)
    f.close()


demoTest1('Total', 1)
demoTest('AH', 1)
demoTest('BJ', 1)
demoTest('CQ', 1)
demoTest('FJ', 1)
demoTest('GD', 1)
demoTest('GS', 1)
demoTest('GX', 1)
demoTest('GZ', 1)
demoTest('HA', 1)
demoTest('HB', 1)
demoTest('HE', 1)
demoTest('HI', 1)
demoTest('HL', 1)
demoTest('HN', 1)
demoTest('JL', 1)
demoTest('JS', 1)
demoTest('JX', 1)
demoTest('LN', 1)
demoTest('NM', 1)
demoTest('NX', 1)
demoTest('QH', 1)
demoTest('SC', 1)
demoTest('SD', 1)
demoTest('SH', 1)
demoTest('SN', 1)
demoTest('SX', 1)
demoTest('TJ', 1)
demoTest('YN', 1)
demoTest('ZJ', 1)
